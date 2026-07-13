"""
Exportiert alle Pipeline-Felddefinitionen als Excel-Datei.
Output: docs/pipeline_io_fields.xlsx
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Daten ────────────────────────────────────────────────────────────────────

FIELDS = [
    # (Agent, Übergang, Feld, Typ, Quelle, Simuliert, Beschreibung)

    # [SYS]
    ("System", "Eingabe → Intake", "raw_input", "str", "Raw Input", "Nein",
     "Ursprünglicher Text: Nachrichtenartikel (Trigger A) oder Nutzeranfrage (Trigger B)"),

    # [I] Intake Agent
    ("Intake Agent", "Intake → Risk Assessment", "relevant", "bool", "LLM", "Nein",
     "Ist die Meldung für die Batterie-Lieferkette relevant?"),
    ("Intake Agent", "Intake → Risk Assessment", "trigger_type", "str (A | B)", "LLM", "Nein",
     "A = automatische News (NewsAPI), B = Nutzeranfrage (Streamlit)"),
    ("Intake Agent", "Intake → Risk Assessment", "material", "str", "LLM", "Nein",
     "Erkanntes Schlüsselmaterial, z.B. 'cobalt', 'lithium'"),
    ("Intake Agent", "Intake → Risk Assessment", "region", "str", "LLM", "Nein",
     "Betroffene Region, z.B. 'Africa/DRC'"),
    ("Intake Agent", "Intake → Risk Assessment", "keywords", "list[str]", "LLM", "Nein",
     "Extrahierte Schlüsselbegriffe aus dem Nachrichtentext"),
    ("Intake Agent", "Intake → Risk Assessment", "filtered_text", "str", "LLM", "Nein",
     "Bereinigter, relevanter Nachrichtentext"),

    # [R] Risk Assessment Agent
    ("Risk Assessment Agent", "Risk Assessment → Network", "severity", "int (1–5)", "LLM", "Nein",
     "Schweregrad: 1=gering, 5=kritisch. LLM-Bewertung aus dem Nachrichtentext"),
    ("Risk Assessment Agent", "Risk Assessment → Network", "risk_type", "str", "LLM", "Nein",
     "Risikoart: supply_disruption | price_volatility | regulatory | logistics | weather"),
    ("Risk Assessment Agent", "Risk Assessment → Network", "affected_material", "str", "LLM", "Nein",
     "Betroffenes Material (bestätigt oder vom LLM präzisiert)"),
    ("Risk Assessment Agent", "Risk Assessment → Network", "affected_region", "str", "LLM", "Nein",
     "Betroffene Region (bestätigt oder vom LLM präzisiert)"),
    ("Risk Assessment Agent", "Risk Assessment → Network", "reason", "str", "LLM", "Nein",
     "LLM-Begründung der Severity-Einschätzung (Pflichtfeld für Validation Agent)"),

    # [N] Network Agent
    ("Network Agent", "Network → Data Retrieval", "affected_nodes", "list[Node]", "Deterministic", "Nein",
     "Betroffene Facilities: MaterialMatch=True AND RegionMatch=True"),
    ("Network Agent", "Network → Data Retrieval", "alt_nodes", "list[Node]", "Deterministic", "Nein",
     "Unbetroffene Alternativen: gleiches Material, andere Region"),
    ("Network Agent", "Network → Data Retrieval", "tier_weights", "dict[str, float]", "Deterministic", "Nein",
     "TierWeight je Facility-ID: Upstream=1.0, BGM=0.8, Cell=0.5, Downstream=0.3"),
    ("Network Agent", "Network → Data Retrieval", "downstream_fanout", "dict[str, int]", "Deterministic", "Nein",
     "Anzahl nachgelagerter Facilities je Upstream-Node (für Handlungsempfehlung)"),

    # Node-Felder (Bestandteil von affected_nodes / alt_nodes)
    ("Network Agent", "Node-Felder", "id", "str", "NAATBatt", "Nein",
     "Facility-ID aus NAATBatt, z.B. '1002'"),
    ("Network Agent", "Node-Felder", "company", "str", "NAATBatt", "Nein",
     "Unternehmensname"),
    ("Network Agent", "Node-Felder", "facility_name", "str", "NAATBatt", "Nein",
     "Name der Produktionsstätte"),
    ("Network Agent", "Node-Felder", "segment", "str", "NAATBatt", "Nein",
     "Upstream | Midstream-BGM | Midstream-Cell | Downstream"),
    ("Network Agent", "Node-Felder", "material_keywords", "str", "NAATBatt", "Nein",
     "z.B. 'cobalt', 'lithium' — bei Downstream-Fabriken leer"),
    ("Network Agent", "Node-Felder", "country / state / city", "str", "NAATBatt", "Nein",
     "Standortangaben der Facility"),
    ("Network Agent", "Node-Felder", "latitude / longitude", "float", "NAATBatt", "Nein",
     "Koordinaten für Streamlit-Kartenmarkierung"),
    ("Network Agent", "Node-Felder", "production_capacity_raw", "str", "NAATBatt / simuliert", "Teilweise",
     "Rohwert als String, z.B. '12000 MT/yr'; 'nan' wenn unbekannt. Quelle: capacity_source"),
    ("Network Agent", "Node-Felder", "capacity_source", "str", "NAATBatt", "Nein",
     "'naatbatt' = Originalwert | 'simulated' = geschätzt"),
    ("Network Agent", "Node-Felder", "single_source_flag", "bool", "Simuliert", "Ja",
     "[simuliert] True wenn ≤2 kommerzielle Anbieter für dieses Material in Nordamerika"),
    ("Network Agent", "Node-Felder", "import_dependency", "bool", "Simuliert", "Ja",
     "[simuliert] True wenn Material importabhängig (nach Materialtyp gesetzt)"),
    ("Network Agent", "Node-Felder", "import_origin_region", "str", "Simuliert", "Ja",
     "[simuliert] z.B. 'Africa/DRC', 'South America / Australia'"),
    ("Network Agent", "Node-Felder", "lead_time_weeks", "int", "Simuliert", "Ja",
     "[simuliert] Vorlaufzeit: Upstream=12W, Midstream-BGM=8W, Midstream-Cell=6W, Downstream=4W"),

    # [D] Data Retrieval Agent
    ("Data Retrieval Agent", "Data Retrieval → Synthesis", "facility_data", "dict[str, FacilityData]", "Deterministic", "Nein",
     "Aufbereitete Kapazitätsdaten je Facility-ID"),
    ("Data Retrieval Agent", "Data Retrieval → Synthesis", "betroffene_kapazitaet_pct", "float", "Deterministic", "Nein",
     "Σ betroffene Kapazität / Σ Gesamtkapazität × 100"),
    ("Data Retrieval Agent", "Data Retrieval → Synthesis", "alternative_kapazitaet_pct", "float", "Deterministic", "Nein",
     "Σ Alternativkapazität / Σ Gesamtkapazität × 100"),

    # FacilityData-Felder
    ("Data Retrieval Agent", "FacilityData-Felder", "capacity", "float", "NAATBatt / simuliert", "Teilweise",
     "Produktionskapazität in MT/yr (0.0 wenn unbekannt)"),
    ("Data Retrieval Agent", "FacilityData-Felder", "capacity_source", "str", "NAATBatt", "Nein",
     "'naatbatt' | 'simulated'"),
    ("Data Retrieval Agent", "FacilityData-Felder", "single_source", "bool", "Simuliert", "Ja",
     "Aus Node.single_source_flag"),
    ("Data Retrieval Agent", "FacilityData-Felder", "import_dep", "bool", "Simuliert", "Ja",
     "Aus Node.import_dependency"),
    ("Data Retrieval Agent", "FacilityData-Felder", "lead_time_norm", "float (0–1)", "Simuliert", "Ja",
     "lead_time_weeks / 12.0"),
    ("Data Retrieval Agent", "FacilityData-Felder", "capacity_share", "float (0–1)", "Deterministic", "Nein",
     "Kapazität dieser Facility / Σ alle Kapazitäten desselben Materials"),

    # [S] Synthesis Agent
    ("Synthesis Agent", "Synthesis → Validation", "risk_report", "str", "LLM", "Nein",
     "Vollständiger Risikobericht mit Quellenangaben"),
    ("Synthesis Agent", "Synthesis → Validation", "top3_facilities", "list[Facility]", "Deterministic", "Nein",
     "Top-3 Facilities nach RiskScore (absteigend)"),
    ("Synthesis Agent", "Synthesis → Validation", "risk_scores", "dict[str, float]", "Deterministic", "Nein",
     "RiskScore (0–100) je Facility-ID"),
    ("Synthesis Agent", "Synthesis → Validation", "global_metrics", "dict", "Deterministic", "Nein",
     "BetroffeneKapazität%, AlternativeKapazität% für Report-Header"),

    # Facility-Felder (Bestandteil von top3_facilities)
    ("Synthesis Agent", "Facility-Felder", "id / company / facility_name / segment", "str", "NAATBatt", "Nein",
     "Basisangaben der Facility"),
    ("Synthesis Agent", "Facility-Felder", "country / state / latitude / longitude", "str / float", "NAATBatt", "Nein",
     "Standortangaben für Streamlit-Kartenmarkierung"),
    ("Synthesis Agent", "Facility-Felder", "risk_score", "float", "Deterministic", "Nein",
     "Rohwert: Severity × TierWeight × Vulnerability × (1 − ResilienceDiscount)"),
    ("Synthesis Agent", "Facility-Felder", "risk_score_normalized", "float (0–100)", "Deterministic", "Nein",
     "risk_score / 5.0 × 100"),
    ("Synthesis Agent", "Facility-Felder", "tier_weight", "float", "Deterministic", "Nein",
     "1.0 / 0.8 / 0.5 / 0.3 je nach Segment"),
    ("Synthesis Agent", "Facility-Felder", "vulnerability", "float (0–1)", "Deterministic", "Nein",
     "0.30×ImportDep + 0.30×SingleSource + 0.25×CapacityShare + 0.15×LeadTime"),
    ("Synthesis Agent", "Facility-Felder", "resilience_discount", "float (0–0.5)", "Deterministic", "Nein",
     "min(AltCapacityRatio / 2, 0.5)"),

    # [V] Validation Agent
    ("Validation Agent", "Validation → Pipeline", "valid", "bool", "LLM", "Nein",
     "Besteht der Bericht alle Prüfungen?"),
    ("Validation Agent", "Validation → Pipeline", "failure_type", "str | None", "LLM", "Nein",
     "'minor' (Retry Synthesis) | 'severe' (Neustart Risk Assessment) | None (valide)"),
    ("Validation Agent", "Validation → Pipeline", "issues", "list[str]", "LLM", "Nein",
     "Liste konkreter Beanstandungen (leer wenn valid=True)"),
    ("Validation Agent", "Validation → Pipeline", "iteration", "int", "Deterministic", "Nein",
     "Iterationszähler; Pipeline bricht bei iteration > 2 ab"),
]

# ── Farben ───────────────────────────────────────────────────────────────────

AGENT_COLORS = {
    "System":                "#f5f5f5",
    "Intake Agent":          "#dae8fc",
    "Risk Assessment Agent": "#d5e8d4",
    "Network Agent":         "#d5e8d4",
    "Data Retrieval Agent":  "#d5e8d4",
    "Synthesis Agent":       "#d5e8d4",
    "Validation Agent":      "#ffe6cc",
}

SIM_COLOR    = "FFD7E8"  # rosa für simulierte Felder
HEADER_COLOR = "37474F"  # dunkelgrau für Header

# ── Excel aufbauen ────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Pipeline IO Felder"

# Header
headers = ["Agent", "Übergang", "Feld", "Typ", "Quelle", "Simuliert?", "Beschreibung"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

ws.row_dimensions[1].height = 28

# Daten eintragen
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for row_idx, (agent, transition, field, typ, source, simulated, desc) in enumerate(FIELDS, 2):
    values = [agent, transition, field, typ, source, simulated, desc]
    agent_hex = AGENT_COLORS.get(agent, "FFFFFF").lstrip("#")

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        if simulated == "Ja":
            cell.fill = PatternFill("solid", fgColor=SIM_COLOR)
        elif col <= 2:
            cell.fill = PatternFill("solid", fgColor=agent_hex)

    ws.row_dimensions[row_idx].height = 30

# Spaltenbreiten
col_widths = [22, 28, 32, 22, 16, 12, 60]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# Einfrieren der Headerzeile
ws.freeze_panes = "A2"

# Legende (zweites Sheet)
ws2 = wb.create_sheet("Legende")
legend = [
    ("Farbe", "Bedeutung"),
    ("Blau (#dae8fc)", "Intake Agent"),
    ("Grün (#d5e8d4)", "Risk Assessment / Network / Data Retrieval / Synthesis Agent"),
    ("Orange (#ffe6cc)", "Validation Agent"),
    ("Rosa", "Simuliertes Feld — Wert nicht aus NAATBatt, sondern manuell gesetzt"),
    ("", ""),
    ("Quelle", "Bedeutung"),
    ("LLM", "Wert wird vom LLM generiert (nicht deterministisch)"),
    ("Deterministic", "Wert wird durch Algorithmus / Arithmetik berechnet"),
    ("NAATBatt", "Wert stammt direkt aus der NAATBatt-Datenbank"),
    ("Simuliert", "Wert manuell gesetzt (kein reales Quelldatum verfügbar)"),
    ("NAATBatt / simuliert", "Teils real, teils simuliert — siehe capacity_source-Feld"),
]
for r, (a, b) in enumerate(legend, 1):
    ws2.cell(row=r, column=1, value=a).font = Font(bold=(r in [1, 7]))
    ws2.cell(row=r, column=2, value=b)
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 70

out = "/Users/yut/Projekt/sc_risk_system/docs/pipeline_io_fields.xlsx"
wb.save(out)
print(f"Gespeichert: {out}")
print(f"Zeilen: {len(FIELDS)} Felder")
